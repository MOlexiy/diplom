import tkinter

import matplotlib
import numpy as np
from datetime import datetime, timedelta
import calendar
import random
from tkinter import *
from tkinter.ttk import Combobox
from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.api import ExponentialSmoothing, SimpleExpSmoothing, Holt
import pmdarima as pm
from sklearn.metrics import mean_squared_error
from math import sqrt
from statsmodels.tsa.arima.model import ARIMAResults
import imgkit
import PIL
import seaborn as sns
import pandas as pd
import matplotlib.pyplot as plt
import csv
from sklearn.preprocessing import PolynomialFeatures
from bokeh.plotting import figure, show, output_file
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression
matplotlib.style.use('ggplot')

window = Tk()

wb2020_2021 = load_workbook('./1628665048603762.xlsx')
wb2021_2022 = load_workbook('./1628611575213049.xlsx')

sheet_1 = wb2020_2021.worksheets[0]
sheet_2 = wb2021_2022.worksheets[0]


year1 = 2020
year2 = year1 + 1
year3 = year2 + 1


def GetDayInMonth(year, month):
    day = 0
    coef = 0
    if year % 4 == 0:
        coef = 1
    else:
        coef = 0
    if month == 1 or month == 3 or month == 5 or month == 7 or month == 8 or month == 10 or month == 12:
        day = 31
    elif month == 2:
        day = 28 + coef
    else:
        day = 30
    return day


def CreateEmptyList(year):
    list = [[]]
    for month in range(1, 13):
        List1 = [[]]
        days = GetDayInMonth(year, month)
        for day in range(1, days+1):
            List2 = []
            for hour in range(1, 25):
                List3 = [hour]
                List2.extend(List3)
            List1.append(List2)
        list.append(List1)
    return list


def CreateTestList(List):
    row = 7
    percent = 0
    for month in range(1, 13):
        days = GetDayInMonth(2020, month)
        for day in range(1, days+1):
            for hour in range(0, 24):
                if sheet_1[row][6].value == hour:
                    List[month][day][hour] = sheet_1[row][8].value
                    row += 1
                else:
                    List[month][day][hour] = random.randint(0, 20)
        percent += 4
        print(percent, '%')
    print('end read exel')
    return List


def CreatePredictList(List):
    row = 7
    percent = 52
    for month in range(1, 13):
        days = GetDayInMonth(2021, month)
        for day in range(1, days + 1):
            for hour in range(0, 24):
                if sheet_2[row][5].value == hour:  # Так как ексель с пропущенными данными, то мы их добавляем рандомайзером от 1 до 10
                    List[month][day][hour] = sheet_2[row][7].value
                    row += 1
                else:
                    List[month][day][hour] = random.randint(0, 20)
        percent += 4
        print(percent, '%')
    print('end read second exel')
    return List


checkList = CreateEmptyList(2020)
ListLavina = CreateTestList(checkList)
checkList = CreateEmptyList(2021)
ListPredictLavina = CreatePredictList(checkList)


def PrintGraphics(x, y, y_pred, text, rmse):
    rmse = str(rmse)
    method = ""
    if text == 0:
        method = "Linear method"
    elif text == 1:
        method = "Polynomial method"
    elif text == 2:
        method = "SARIMA method"
    elif text == 3:
        method = "Holt-Winters method"
    p = figure(title="Actual vs Predict and method: " + method + " and RMSE: " + rmse, width=1350, height=900)
    p.title.align = 'center'
    p.circle(x, y_pred)
    p.line(x, y, legend_label='Actual', line_width=3, line_alpha=0.4)
    p.circle(x, y_pred, color="red")
    p.line(x, y_pred, color="red", legend_label='Predict', line_width=3, line_alpha=0.4)
    p.xaxis.axis_label = 'day in month by hour (day * hour)'
    p.yaxis.axis_label = 'Visitors'
    show(p)


def GetDeltaInMonth(month, year, needDay):
    day = 0
    date = datetime(year, month, 1)
    start_day_in_month = calendar.monthrange(date.year, date.month)[0]
    if start_day_in_month != needDay:
        day = needDay - start_day_in_month
    return day


def GetArimaArray(Aday, Amonth):
    ListForArima = []
    for years in range(2020, 2021):
        if years == 2020:
            for month in range(Amonth, 13):
                days = GetDayInMonth(years, month)
                if month == Amonth:
                    for day in range(Aday, days + 1):
                        for hour in range(0, 24):
                            ListForArima.append(ListLavina[month][day][hour])
                elif month > Amonth:
                    for day in range(1, days + 1):
                        for hour in range(0, 24):
                            ListForArima.append(ListLavina[month][day][hour])

        elif years == 2021:
            for month in range(1, Amonth + 1):
                if month < Amonth:
                    days = GetDayInMonth(years, month)
                    for day in range(1, days + 1):
                        for hour in range(0, 24):
                            ListForArima.append(ListPredictLavina[month][day][hour])
                elif month == Amonth:
                    for day in range(1, Aday):
                        for hour in range(0, 24):
                            ListForArima.append(ListPredictLavina[month][day][hour])
    return ListForArima


def Regress(month, countDays, numberWeek, lvl0, lvl1, lvl2, text3):  # TODO пофіксити аріму
    pf = PolynomialFeatures(degree=5, include_bias=False)
    coefForYLMMD = getYellowOrRed(lvl0)
    coefForYYMLD = getYellowOrRed(lvl1)
    coefForYYMMD = getYellowOrRed(lvl2)
    ListDay = []  # список для графика дней
    ListHour = []
    ListDataActual = []
    ListDataPredict = []
    ListDataLastDay = []
    ListDataLastWeekDay = []
    ListSecondDataTrainingTemp = []
    ListDataTemp = []
    ListDataHalfTemp = []

    firstDayInPredictMonth = calendar.monthrange(2021, month)[0]
    checkDay = firstDayInPredictMonth
    print(coefForYLMMD)
    firstDelta = GetDeltaInMonth(month-1, 2021, firstDayInPredictMonth)
    countDayInLastMonth = calendar.monthrange(2021, month-1)[1]
    secondDelta = GetDeltaInMonth(month, 2020, firstDayInPredictMonth)
    countDayInLastYearMinosMonth = calendar.monthrange(2020, month-1)[1]
    startBlock = 1 + 7 * numberWeek
    endBlock = countDays + 7 * numberWeek

    TempListForArima = GetArimaArray(startBlock, month)

    for days in range(1, countDays * 24 + 1):
        ListDay.append(days)
    for hour in range(0, 24):
        ListHour.append(hour)
        ListHour.append(hour)
        ListHour.append(hour)
        ListHour.append(hour)
    x = np.array(ListHour).reshape((-1, 1))
    pf.fit(x)

    for days in range(startBlock, endBlock + 1):
        countForFList = firstDelta
        countForSList = secondDelta
        ListDataTemp.clear()
        ListSecondDataTrainingTemp.clear()
        index = 0
        for hour in range(0, 24):  # TODO refactor this
            tempMonth = month
            if countForFList < 0:
                tempMonth -= 2
            else:
                tempMonth -= 1
            ListSecondDataTrainingTemp.append(ListPredictLavina[tempMonth][days - firstDelta + 1][hour])
            # ListDataHalfTemp.append(ListPredictLavina[tempMonth][days - firstDelta][hour])
            ListDataActual.append(ListPredictLavina[month][days][hour])
            if 1 <= checkDay <= 4:
                if countForSList < 0:
                    ListDataTemp.append(int(ListLavina[month-1][countDayInLastYearMinosMonth + countForSList][hour] / coefForYLMMD[1]))
                else:
                    ListDataTemp.append(int(ListLavina[month][days - secondDelta + 1][hour] / coefForYLMMD[1]))
                ListDataTemp.append(int(ListSecondDataTrainingTemp[index] / coefForYYMLD[1]))
            elif checkDay == 0:
                if countForSList < 0:
                    if countDays == 2:
                        ListDataTemp.append(int(ListLavina[month-1][countDayInLastYearMinosMonth + countForSList][hour] / 0.1))
                    else:
                        ListDataTemp.append(
                            int(ListLavina[month - 1][countDayInLastYearMinosMonth + countForSList][hour] / 0.13))
                else:
                    ListDataTemp.append(int(ListLavina[month][days - secondDelta + 1][hour] / random.uniform(coefForYYMLD[0], coefForYYMLD[1])))
                ListDataTemp.append(int(ListSecondDataTrainingTemp[index] / random.uniform(coefForYYMLD[0], coefForYYMLD[1])))
            else:
                if countForSList < 0:
                    ListDataTemp.append(int(ListLavina[month-1][countDayInLastYearMinosMonth + countForSList][hour] / coefForYLMMD[0]))
                else:
                    print(days - secondDelta)
                    ListDataTemp.append(int(ListLavina[month][days - secondDelta + 1][hour] / coefForYLMMD[0]))
                ListDataTemp.append(int(ListSecondDataTrainingTemp[index] / coefForYYMLD[0]))
            if startBlock >= 7:
                ListDataLastWeekDay.append(ListPredictLavina[month][days - 7][hour])
                if int(ListDataPredict.__len__()) == 0:
                    ListDataLastDay.append(ListPredictLavina[month][days - 1][hour])
            else:
                tempCount = countDays - 1
                ListDataLastWeekDay.append(ListPredictLavina[month - 1][countDayInLastMonth - tempCount][hour])
                if int(ListDataPredict.__len__()) == 0:
                    ListDataLastDay.append(ListPredictLavina[month - 1][countDayInLastMonth][hour])
                tempCount -= 1
            if int(ListDataPredict.__len__()) > 0:
                tempCoef = int(ListDataPredict.__len__())
                ListDataLastDay.append(int(ListDataPredict[int(hour * 2 + tempCoef - 48)]))
            ListDataTemp.append(ListDataLastDay[index])
            ListDataTemp.append(ListDataLastWeekDay[index])
            index += 1
        checkDay += 1
        if checkDay == 7:
            checkDay = 0
        countForFList += 1
        countForSList += 1
        if text3 == 0:
            model = LinearRegression().fit(x, ListDataTemp)
            y_pred = model.predict(x)
            ListDataPredict.extend(y_pred * coefForYYMMD[0])
        elif text3 == 1:
            x_ = pf.transform(x)
            model = LinearRegression().fit(x_, ListDataTemp)
            y_pred = abs(model.predict(x_))
            ListDataPredict.extend(y_pred * coefForYYMMD[0])
        elif text3 == 3:
            model = ExponentialSmoothing(np.asarray(ListDataTemp), seasonal_periods=7, trend='add', seasonal='add').fit()
            y_pred = abs(model.forecast(24))
            ListDataPredict.extend(y_pred * coefForYYMMD[0])
    if text3 == 2:
        # model = pm.auto_arima(TempListForArima, start_p=1, start_q=1,  # find best settings for arima method
        #                       test='adf',
        #                       max_p=3, max_q=3,
        #                       m=12,
        #                       d=None,
        #                       seasonal=True,
        #                       start_P=0,
        #                       D=1,
        #                       trace=True,
        #                       error_action='ignore',
        #                       suppress_warnings=True,
        #                       stepwise=True)
        # print(model.summary())
        model = ARIMA(TempListForArima, order=(2, 1, 4), seasonal_order=(0, 1, 2, 24)).fit()
        yhat = abs(model.predict(24*countDays, alpha=0.05, dynamic=True))
        ListDataPredict.extend(yhat * coefForYYMMD[0])
    if 0 <= text3 <= 1:
        ListDataPredict = getMathAverage(ListDataPredict)
    print(ListDataActual.__len__())
    print(ListDataPredict.__len__())
    rmse = sqrt(mean_squared_error(ListDataActual, ListDataPredict))  # TODo fix rmse like <100%
    # print('Test RMSE: %.3f' % rmse)
    # writeArray(ListDataPredict, ListDataActual, month)
    PrintGraphics(ListDay, ListDataActual, ListDataPredict, text3, rmse)


def getMathAverage(array):
    arraytemp = []
    for lenar in range(0, int(array.__len__() / 4)):
        average = int((array[(lenar * 4)] + array[(lenar * 4 + 1)] + array[(lenar * 4 + 2)] + array[(lenar * 4 + 3)]) / 4)
        arraytemp.append(average)
    return arraytemp


def writeArray(array, arraylavina, month):
    day = GetDayInMonth(year1, month)
    array2 = []
    array3 = []
    for days in range(1, day+1):
        for hour in range(0, 24):
            array2.append(hour)
            array3.append(days)
    df = pd.DataFrame({'Day': array3, 'Hour': array2, 'Visitors': array, 'Lavina': arraylavina})
    df.to_excel('./predict.xlsx')


def getYellowOrRed(lvl):  # TODO описати в вигляді тексту в диплом метод.
    coef = []
    if lvl == 1:
        coef = [0.5, 0.6]
        # coef = random.uniform(0.45, 0.75)
    elif lvl == 2:
        coef = [0.07, 0.13]
        # coef = random.uniform(0.05, 0.17)
    elif lvl == 0:
        coef = [1, 1]
    return coef


def getLvl(text):
    lvl = 0
    if text == 0:
        lvl = 0
    elif text == 1:
        lvl = 1
    elif text == 2:
        lvl = 2
    return lvl


def clicked():
    enterMonth = int(combo0.get())
    countDays = int(combo1.get())
    numberWeek = int(combo2.get()) - 1
    text0 = int(selected0.get())
    text1 = int(selected1.get())
    text2 = int(selected2.get())
    text3 = int(selected3.get())
    lvl0 = getLvl(text0)
    lvl1 = getLvl(text1)
    lvl2 = getLvl(text2)
    Regress(enterMonth, countDays, numberWeek, lvl0, lvl1, lvl2, text3)


# Display menu
window.title("Interface of program")
window.geometry('650x330')
lbl0 = Label(window, text="Enter a month number")
lbl0.grid(column=0, row=0)
combo0 = Combobox(window)
combo0['values'] = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12)
combo0.current(0)
combo0.grid(column=1, row=0)

lbl1 = Label(window, text="Enter the number of days and week of reference")
lbl1.grid(column=0, row=1)
combo1 = Combobox(window)
combo1['values'] = (5, 7)
combo1.current(0)
combo1.grid(column=1, row=1)
combo2 = Combobox(window)
combo2['values'] = (1, 2, 3, 4)
combo2.current(0)
combo2.grid(column=2, row=1)

lbl2 = Label(window, text="Add settings of restrictions for prediction")
lbl2.grid(column=0, row=2)
selected0 = IntVar()
rad0 = Radiobutton(window, text='NON', value=0, variable=selected0)
rad1 = Radiobutton(window, text='Yellow', value=1, variable=selected0)
rad2 = Radiobutton(window, text='Red', value=2, variable=selected0)
rad0.grid(column=0, row=3)
rad1.grid(column=1, row=3)
rad2.grid(column=2, row=3)

lbl3 = Label(window, text="Add constraint settings for the past week")
lbl3.grid(column=0, row=4)
selected1 = IntVar()
radb0 = Radiobutton(window, text='NON', value=0, variable=selected1)
radb1 = Radiobutton(window, text='Yellow', value=1, variable=selected1)
radb2 = Radiobutton(window, text='Red', value=2, variable=selected1)
radb0.grid(column=0, row=5)
radb1.grid(column=1, row=5)
radb2.grid(column=2, row=5)

# settings / add constraint settings for this week past year. / like pred 2021 04 (13-18 d), past year wb 2020 04 (13-18 d)
lbl4 = Label(window, text="Add constraint settings for this week past year")
lbl4.grid(column=0, row=6)
selected2 = IntVar()
radiobutton0 = Radiobutton(window, text='NON', value=0, variable=selected2)
radiobutton1 = Radiobutton(window, text='Yellow', value=1, variable=selected2)
radiobutton2 = Radiobutton(window, text='Red', value=2, variable=selected2)
radiobutton0.grid(column=0, row=7)
radiobutton1.grid(column=1, row=7)
radiobutton2.grid(column=2, row=7)

lbl5 = Label(window, text="Enter the method of predict")
lbl5.grid(column=0, row=8)
selected3 = IntVar()
predictBut0 = Radiobutton(window, text='Linear', value=0, variable=selected3)
predictBut1 = Radiobutton(window, text='Polynomial', value=1, variable=selected3)
predictBut2 = Radiobutton(window, text='ARIMA', value=2, variable=selected3)
predictBut3 = Radiobutton(window, text='Holt-Winters', value=3, variable=selected3)
predictBut0.grid(column=0, row=9)
predictBut1.grid(column=1, row=9)
predictBut2.grid(column=2, row=9)
predictBut3.grid(column=3, row=9)

btn = Button(window, text="Enter", command=clicked)
btn.grid(column=0, row=10)

# end Display menu
window.mainloop()